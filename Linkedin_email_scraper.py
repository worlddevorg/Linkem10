import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
import re
import random
import openpyxl

workbook_path = None  # Initialize the global variable
try:
    with open('last_processed_row.txt', 'r') as file:
        last_processed_row = int(file.read())
except FileNotFoundError:
    last_processed_row = 2
def solve_captcha():
    print("Bypassing Captcha please wait...")
    time.sleep(200)
    try:
        captchasubmit = driver.find_element('xpath', '//*[@id="recaptcha-demo-submit"]').click()
    except Exception as e:
        print("Failed to bypass captcha:", e)

def import_sheet():
    global workbook_path
    workbook_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not workbook_path:
        return

def extract_data(tag):
    global workbook_path
    if not workbook_path:
        messagebox.showerror("Error", "Please import a workbook first.")
        return
    global driver
    global last_processed_row
    options = webdriver.ChromeOptions()
    options.binary_location = "/bin/brave-browser-stable"
    driver = webdriver.Chrome(options=options)
    actions = ActionChains(driver)
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    extracted_workbook = openpyxl.Workbook()
    extracted_sheet = extracted_workbook.active
    extracted_sheet['A1'] = 'Name'
    extracted_sheet['B1'] = 'Email'
    extracted_sheet['C1'] = 'Designation'
    extracted_row = 2
    for row in range(last_processed_row, sheet.max_row + 1):
        company_name = sheet.cell(row=row, column=1).value
        if not company_name:
            continue
        for tag in tags:
            driver.get('https://www.startpage.com/')
            g_search = driver.find_element('xpath', '//*[@id="q"]')
            selected_domain = email_domain_combobox.get()
            g_search.send_keys(f'site:linkedin.com/ "{company_name}" "{tag}" "{selected_domain}"')
            g_search.send_keys(Keys.ENTER)
            time.sleep(random.randint(1,4))
            matches_found = 0
            max_matches = 3
            while matches_found < max_matches:
                matches = []
                for div in driver.find_elements('xpath', '//div[@class="result css-z73qjy"]'):
                    # print(type(tag),type(div),tag,div)
                    if str(tag.lower()) in str(div.text.lower()):
                        print(div.text)
                        lines = div.text.split('\n')
                        if len(lines) > 1:
                            second_line = lines[1]
                            parts = second_line.split('-')
                            print(parts)
                            if len(parts) > 1:
                                job_role=parts[1].strip()
                            elif len(parts) > 0:
                                job_role=str(tag)
                        div_text = re.sub(r'https?://\S+', '', div.text)
                        name = ' '.join(div_text.split()[:2])
                        email_match = re.search(r'\b[\w\.-]+@gmail\.com\b', div_text)
                        if email_match:
                            email = email_match.group()
                        else:
                            email = None
                        matches.append((name, email,job_role))
                    else:
                        pass
                for match in matches:
                    # print(match)
                    if match[0] is not None and match[1] is not None:
                        if "+" not in match[0]:
                            email = match[1].split(",")[0].strip()
                            extracted_sheet.cell(row=extracted_row, column=1).value = match[0]
                            extracted_sheet.cell(row=extracted_row, column=2).value = email
                            extracted_sheet.cell(row=extracted_row, column=3).value = f"{match[2]} at {company_name}"
                            extracted_row += 1
                            # matches_found += 1
                            if matches_found >= max_matches:
                                break
                try:
                    clickbtn = driver.find_element('xpath', '//button[contains(text(), "Next")]').click()
                    time.sleep(2)
                except:
                    pass
                matches_found += 1
        extracted_workbook.save("Linkdein_emails.xlsx")
        last_processed_row+=1
        with open('last_processed_row.txt', 'w') as file:
            file.write(str(last_processed_row))
    driver.quit()
    messagebox.showinfo("Extraction Complete", "Data extraction process has been completed successfully.")

def add_tags():
    tag = tags_entry.get()
    if tag:
        tags_listbox.insert(tk.END, tag)
        tags_entry.delete(0, tk.END)

def remove_tags():
    selected_index = tags_listbox.curselection()
    if selected_index:
        tags_listbox.delete(selected_index)

def start_scraping():
    global tags
    tags = tags_listbox.get(0, tk.END)
    for tag in tags:
        extract_data(tag)

def main():
    root = tk.Tk()
    root.title("Caliber Inc Developer - Waqar Ali Abbas")
    root.geometry("570x400")
    root.resizable(False, False)
    
    style = ttk.Style()
    style.theme_use("clam")  # Use a modern theme
    
    tags_frame = ttk.Frame(root)
    tags_frame.pack(pady=10, padx=10, fill="both", expand=True)
    
    tags_label = ttk.Label(tags_frame, text="Designation:")
    tags_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
    
    global tags_entry
    tags_entry = ttk.Entry(tags_frame, width=25)
    tags_entry.grid(row=0, column=1, padx=5, pady=5)
    
    add_button = ttk.Button(tags_frame, text="Add", command=add_tags)
    add_button.grid(row=0, column=2, padx=5, pady=5)
    
    remove_button = ttk.Button(tags_frame, text="Remove", command=remove_tags)
    remove_button.grid(row=0, column=3, padx=5, pady=5)
    
    global tags_listbox
    tags_listbox = tk.Listbox(root, width=35, height=6)
    tags_listbox.pack(pady=5, padx=10, fill="both", expand=True)
    
    import_button = ttk.Button(root, text="Import Workbook", command=import_sheet)
    import_button.pack(pady=10, padx=10, fill="x")
    
    email_domain_label = ttk.Label(root, text="Select Email Domain")
    email_domain_label.pack(pady=5, padx=10, anchor="w")
    
    global email_domain_combobox
    email_domain_combobox = ttk.Combobox(root, values=["@gmail.com", "@yahoo.com", "@hotmail.com"])
    email_domain_combobox.pack(pady=5, padx=10, fill="x")
    email_domain_combobox.set("@gmail.com")
    
    scraping_button = ttk.Button(root, text="Start Scraping", command=start_scraping)
    scraping_button.pack(pady=10, padx=10, fill="x")
    
    root.mainloop()

if __name__ == "__main__":
    main()
